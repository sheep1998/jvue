# -*- coding: utf-8 -*-
"""
Created on Sat Jul 13 01:31:24 2019

@author: 1535725170
"""
import pandas as pd
import numpy as np
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def txtRead(filename):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\txt\\"+filename
    f = open(path,'r',encoding='utf-8')
    content = f.readlines()
    print(content)
    contentList = []
    for line in content:
        #line = line.replace("\n","")
        line = line.strip()
        contentList.append(line)
    f.close()
    return contentList
  
def txtWrite(filename,contentList):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\txt\\"+filename
    f = open(path,'w',encoding='utf-8')
    for line in contentList:
        f.write(line+'\n')
    f.close()

def csvRead2List(filename):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\csv\\"+filename
    ws = pd.read_csv(path,header=None)
#    print(ws[0:3])
#    print(ws[0:1][2])
#    print(ws[2][0])
    ws = ws.values.tolist()
#    print(len(ws))
#    print(ws[0][2])
    return ws

def csvRead2Df(filename):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\csv\\"+filename
    #ws = pd.read_csv(path,header=None,names=['t1','t2','t3'])
    ws = pd.read_csv(path)
    return ws

def csvWriteDf(filename,df):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\csv\\"+filename
    #不保存列名
    df.to_csv(path,header=0,index=0)
    
def csvWriteList(filename,lst):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\csv\\"+filename
    out = open(path,'w',newline='')
    #out = open(path,'a',newline='')
    csv_write = csv.writer(out,dialect='excel')
    for line in lst:
        csv_write.writerow(line)
    out.close()
    
def csvRead2List2(filename):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\csv\\"+filename
    out = open(path,'r')
    csv_file = csv.reader(out)
    #遍历打印
    return csv_file

def xlsxRead(filename):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\xlsx\\"+filename
    wb = load_workbook(filename=path)
    ws = wb['Sheet 1']
    for row in range(1,10,1):
        print(ws.cell(row=row,column=2).value)
    
def xlsmWriteDf(filename,df):
    path = "C:\\Users\\1535725170\\Documents\\GitHub\\jvue\\xlsx\\"+filename
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df,index=False,header=True):
        ws.append(row)
    wb.save(path)

#lst = csvRead2List("data.csv")
#csvWriteList("data2.csv",lst)
#df = csvRead2Df("data2.csv")
#csvWriteDf("data3.csv",df)
#xlsxRead("data1.xlsx")
#xlsmWriteDf("data2.xlsx",df)

def pdDemo():
    frame=['t1','t2','t3']
    df = pd.DataFrame(columns=frame)
    
    for pos in range(10):
        arr = [2,3,4]
        df.loc[pos] = arr
    return df

#print(pdDemo())
    
class Translator:
    def __init__(self):
        self.a = 1
    
    def s2i(self,s):
        return int(s)
    
    def s2f(self,s):
        return '%.2f' % float(s)
    
    def num2s(self,num):
        return str(num)
    
    def s2low(self,s):
        return s.lower()
    
    def s2up(self,s):
        return s.upper()

    def np2df(self,np):
        df = pd.DataFrame(np)
        return df
    
    def df2np(self,df):
        return df.values
    
    def np2lst(self,npp):
        return npp.tolist()
    
    def lst2np(self,lst):
        return np.array(lst)
